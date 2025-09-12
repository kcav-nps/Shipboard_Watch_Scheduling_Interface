# export_service.py
# -----------------------------------------------------------------------------
# Exports for:
#   1) Personnel Excel with two sheets (Rollup / Formatted)
#   2) Monthly overview Excel (Daily Overview + detail sheets)
#      + NEW: "Leaves (From-To)" sheet that compresses daily leave rows to ranges
#
# Requirements: pandas, openpyxl
# -----------------------------------------------------------------------------

from __future__ import annotations

from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.utils import get_column_letter

from .personnel_service import list_personnel, _is_officer

# --------------------------- Helpers (formatting) ----------------------------

def _autofit_worksheet(ws):
    """Auto-fit column widths based on content length (capped)."""
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

def _freeze_and_fit(ws):
    """Freeze header row, add filter, and auto-fit columns."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit_worksheet(ws)

# ------------------------ 1) Personnel Excel export -------------------------

def export_personnel_excel(outfile: str | Path) -> Path:
    """
    Create an Excel with two sheets:
      - 'Rollup': one column per field (human-friendly English headers)
      - 'Formatted': two blocks (Officers / Warrant & NCO) using naming rule
    """
    df = list_personnel()

    # Ensure columns exist even if empty
    for c in [
        "name","rank","specialty","duty",
        "primary_shift","alt_shift","at_sea_shift",
        "height","weight","registry_number","address","phone",
        "marital_status","children","pye_expiration","notes"
    ]:
        if c not in df.columns:
            df[c] = ""

    # English headers
    headers = {
        "rank":"Rank",
        "name":"Full Name",
        "specialty":"Specialty",
        "duty":"Duties",
        "primary_shift":"Primary in-port watch",
        "alt_shift":"Alt in-port watch",
        "at_sea_shift":"At-sea watch",
        "height":"Height (cm)",
        "weight":"Weight (kg)",
        "registry_number":"Registry Number",
        "address":"Address",
        "phone":"Phone",
        "marital_status":"Marital status",
        "children":"Children",
        "pye_expiration":"PYE valid until",
        "notes":"Notes",
    }
    df_out = df[list(headers.keys())].rename(columns=headers)

    # Build "Formatted"
    df["__is_officer"] = df["rank"].apply(_is_officer)
    off = df[df["__is_officer"]].copy()
    nco = df[~df["__is_officer"]].copy()

    off_out = pd.DataFrame({
        "Rank": off["rank"],
        "Full Name HN": off["name"].astype(str).str.strip() + " HN" # Assuming ΠΝ meant Hellenic Navy
    })
    nco_out = pd.DataFrame({
        "Rank (Specialty)": nco["rank"].astype(str).str.strip() + " (" + nco["specialty"].astype(str).str.strip() + ")",
        "Full Name": nco["name"].astype(str).str.strip()
    })

    # Write workbook
    outfile = Path(outfile)
    with pd.ExcelWriter(outfile, engine="openpyxl") as xw:
        # Sheet 1: Rollup
        df_out.to_excel(xw, sheet_name="Rollup", index=False)
        ws = xw.book["Rollup"]
        _freeze_and_fit(ws)

        # Sheet 2: Formatted
        off_out.to_excel(xw, sheet_name="Formatted", index=False, startrow=1)
        ws2 = xw.book["Formatted"]
        ws2["A1"] = "Officers"
        _autofit_worksheet(ws2)
        start = ws2.max_row + 2
        ws2[f"A{start}"] = "Warrant & NCOs"
        nco_out.to_excel(xw, sheet_name="Formatted", index=False, startrow=start)
        _autofit_worksheet(ws2)

    return outfile

# ---------------------- 2) Monthly overview Excel export --------------------

LOGS_DIR = Path("logs")

def _load_monthly_csv(name: str, year: int, month: int, cols: list[str]) -> pd.DataFrame:
    """
    Load a monthly CSV from ./logs with canonical columns.
    If the file does not exist, return an empty DataFrame with those columns.
    """
    path = LOGS_DIR / f"{name}_{year:04d}_{month:02d}.csv"
    if path.exists():
        df = pd.read_csv(path, dtype=str).fillna("")
    else:
        df = pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df

def _compress_date_ranges(leaves_df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert daily leave rows to contiguous date ranges per (registry_id, leave_type, comments).
    Input columns (expected): date, registry_id, leave_type, comments
    Output columns: From, To, Registry No., Leave Type, Comments, Days
    """
    if leaves_df.empty:
        return pd.DataFrame(columns=["From","To","Registry No.","Leave Type","Comments","Days"])

    df = leaves_df.copy().fillna("")
    # Ensure required columns
    for c in ["date", "registry_id", "leave_type", "comments"]:
        if c not in df.columns:
            df[c] = ""

    # Parse and sort by (person, type, comments, date)
    df["__dt"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["__dt"]).sort_values(["registry_id", "leave_type", "comments", "__dt"])

    rows = []
    for (reg, ltype, comm), g in df.groupby(["registry_id", "leave_type", "comments"], dropna=False):
        g = g.sort_values("__dt")
        start = prev = None
        count = 0
        for d in g["__dt"]:
            if start is None:
                start = prev = d
                count = 1
                continue
            # Extend contiguous range or flush current
            if d == prev + timedelta(days=1):
                prev = d
                count += 1
            else:
                rows.append({
                    "From": start.date().isoformat(),
                    "To": prev.date().isoformat(),
                    "Registry No.": reg,
                    "Leave Type": ltype,
                    "Comments": comm,
                    "Days": count
                })
                start = prev = d
                count = 1
        # Flush the last open range
        if start is not None:
            rows.append({
                "From": start.date().isoformat(),
                "To": prev.date().isoformat(),
                "Registry No.": reg,
                "Leave Type": ltype,
                "Comments": comm,
                "Days": count
            })

    out = pd.DataFrame(rows, columns=["From","To","Registry No.","Leave Type","Comments","Days"])
    return out

def export_monthly_overview(year: int, month: int, outfile: str | Path) -> Path:
    """
    Build a single Excel file with:
      - 'Daily Overview': one row per encountered date with ship status, holiday flag, and counts
      - 'Leaves (From-To)': compressed leave ranges
      - 'Leaves', 'Unavailabilities', 'Preferences', 'Holidays', 'Ship Status' (daily-detail sheets)
    """
    # Load monthly logs
    leaves = _load_monthly_csv("daily_leave", year, month,
                               ["date","registry_id","leave_type","comments"])
    cannot = _load_monthly_csv("daily_cannot", year, month,
                               ["date","registry_id","comments"])
    prefer = _load_monthly_csv("daily_prefer", year, month,
                               ["date","registry_id","comments"])
    hol    = _load_monthly_csv("holidays", year, month,
                               ["date","description"])
    ship   = _load_monthly_csv("ship_status", year, month,
                               ["date","status"])

    # Collect all dates present in any table
    all_dates = set()
    for df in (leaves, cannot, prefer, hol, ship):
        if "date" in df.columns:
            all_dates |= set(df["date"].dropna().astype(str).tolist())

    # If no data at all, create a minimal workbook with an empty overview
    outfile = Path(outfile)
    if not all_dates:
        with pd.ExcelWriter(outfile, engine="openpyxl") as xw:
            pd.DataFrame(columns=[
                "Date","Ship Status","Holiday","Description",
                "# Leaves","# Unavailabilities","# Preferences"
            ]).to_excel(xw, sheet_name="Daily Overview", index=False)
        return outfile

    # Sort dates ascending
    try:
        sorted_dates = sorted(all_dates, key=lambda s: datetime.fromisoformat(s))
    except Exception:
        sorted_dates = sorted(all_dates)

    # Pre-compute counts
    cnt_leave  = leaves.groupby("date").size().to_dict() if not leaves.empty else {}
    cnt_cannot = cannot.groupby("date").size().to_dict() if not cannot.empty else {}
    cnt_prefer = prefer.groupby("date").size().to_dict() if not prefer.empty else {}

    # Maps for ship status & holidays
    ship_map = ship.set_index("date")["status"].to_dict() if not ship.empty else {}
    hol_map  = hol.set_index("date")["description"].to_dict() if not hol.empty else {}

    # Build overview rows
    rows = []
    for d in sorted_dates:
        rows.append({
            "Date": d,
            "Ship Status": ship_map.get(d, ""),
            "Holiday": "Yes" if d in hol_map else "No",
            "Description": hol_map.get(d, ""),
            "# Leaves": cnt_leave.get(d, 0),
            "# Unavailabilities": cnt_cannot.get(d, 0),
            "# Preferences": cnt_prefer.get(d, 0),
        })
    overview = pd.DataFrame(rows)

    # Human-friendly column names for detail sheets
    leaves_out = leaves.rename(columns={
        "date":"Date","registry_id":"Registry No.",
        "leave_type":"Leave Type","comments":"Comments"
    })
    cannot_out = cannot.rename(columns={
        "date":"Date","registry_id":"Registry No.","comments":"Comments"
    })
    prefer_out = prefer.rename(columns={
        "date":"Date","registry_id":"Registry No.","comments":"Comments"
    })
    hol_out = hol.rename(columns={
        "date":"Date","description":"Description"
    })
    ship_out = ship.rename(columns={
        "date":"Date","status":"Ship Status"
    })

    # NEW: compressed "from-to" view for leaves
    leaves_ranges = _compress_date_ranges(leaves)

    # Write workbook
    with pd.ExcelWriter(outfile, engine="openpyxl") as xw:
        overview.to_excel(xw, sheet_name="Daily Overview", index=False)
        _freeze_and_fit(xw.book["Daily Overview"])

        # NEW: Leaves as ranges (From-To)
        leaves_ranges.to_excel(xw, sheet_name="Leaves (From-To)", index=False)
        _freeze_and_fit(xw.book["Leaves (From-To)"])

        # Existing daily-detail sheets
        leaves_out.to_excel(xw, sheet_name="Leaves", index=False)
        _freeze_and_fit(xw.book["Leaves"])

        cannot_out.to_excel(xw, sheet_name="Unavailabilities", index=False)
        _freeze_and_fit(xw.book["Unavailabilities"])

        prefer_out.to_excel(xw, sheet_name="Preferences", index=False)
        _freeze_and_fit(xw.book["Preferences"])

        hol_out.to_excel(xw, sheet_name="Holidays", index=False)
        _freeze_and_fit(xw.book["Holidays"])

        ship_out.to_excel(xw, sheet_name="Ship Status", index=False)
        _freeze_and_fit(xw.book["Ship Status"])

    return outfile

