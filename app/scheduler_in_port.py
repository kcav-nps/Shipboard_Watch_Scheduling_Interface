# app/scheduler_in_port.py (DEBUGGING VERSION)
# -----------------------------------------------------------------------------
# This version contains extra print statements to diagnose the "blank schedule" issue.
# -----------------------------------------------------------------------------

from __future__ import annotations
from .i18n_display_mapping import I18N
from .constants import RANKS, SPECIALTIES, DUTIES

from pathlib import Path
from datetime import datetime
import pandas as pd

from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Project-local helpers and rules
from .scheduling_prep import day_availability, WATCH_TYPES, _seniority_key
from .i18n_display_mapping import I18N
from .constants import RANKS, SPECIALTIES, DUTIES
from .scheduler_rules import (
    MAX_PER_MONTH, DUTY_NEVER, DUTY_WEEKDAY_ONLY,
    is_weekend, is_weekday, is_holiday, weekday_name_gr, two_day_gap_ok
)

# ------------------------- Name formatting helpers ---------------------------

I18N_SCH = I18N(RANKS, SPECIALTIES, DUTIES, ["", "AF","YF","YFM","BYFM","BYF"])

def _is_officer(rank: str) -> bool:
    officers = {
        "Commander","Commander (M)","Lieutenant Commander","Lieutenant Commander (M)",
        "Lieutenant","Lieutenant (M)","Lieutenant (E)",
        "Ensign","Ensign (M)","Ensign (E)"
    }
    return str(rank).strip() in officers

def _display_name(rank: str, specialty: str, name: str) -> str:
    rank = I18N_SCH.to_storage('rank', str(rank).strip())
    specialty = I18N_SCH.to_storage('specialty', str(specialty).strip())
    name = str(name).strip()
    if _is_officer(rank):
        return f"{rank} | {name} HN"
    spec = f" ({specialty})" if specialty else ""
    return f"{rank}{spec} | {name}"

# ------------------------------ Data loading ---------------------------------

def _load_people_map() -> dict:
    people_path = Path("data") / "personnel.csv"
    if not people_path.exists():
        return {}
    people = pd.read_csv(people_path, dtype=str, encoding='utf-8-sig').fillna("")
    out = {}
    for _, r in people.iterrows():
        key = str(r.get("registry_number", "")).strip()
        if key:
            out[key] = r.to_dict()
    return out

# ---------------------------- Pools & ordering -------------------------------

def _primary_pool(info: dict, watch: str, people_map: dict, af_mode: bool = False) -> list[dict]:
    base = info.get("pools", {}).get(watch, [])
    out = []
    print(f"DEBUG: Filtering primary pool for watch '{watch}'. Initial size: {len(base)}.")
    for p in base:
        rid = str(p.get("registry_id", "")).strip()
        row = people_map.get(rid)
        if row is None:
            continue
        duty = str(row.get("duty", "")).strip()
        if duty in DUTY_NEVER:
            continue
        
        primary_shift = str(row.get("primary_shift",'')).strip()
        if primary_shift != watch:
            # print(f"  - DEBUG: Excluding {p.get('name')} because primary_shift is '{primary_shift}' not '{watch}'.")
            continue
        
        q = dict(p)
        if af_mode:
            q["duty"] = duty
        out.append(q)
    print(f"DEBUG: Final primary pool size for '{watch}': {len(out)}.")
    return out

def _sort_af_youngest_first(pool: list[dict]) -> list[dict]:
    return sorted(pool, key=lambda p: (_seniority_key(p["rank"]), str(p["name"])), reverse=True)

def _sort_fair_non_af(pool: list[dict], counters: dict) -> list[dict]:
    def key(p):
        rid = p["registry_id"]
        tot = counters.get(rid, {}).get("total", 0)
        return (tot, _seniority_key(p["rank"]), str(p["name"]))
    return sorted(pool, key=key)

# ------------------------------ Date constraints -----------------------------

def _is_holiday_like(date_iso: str) -> bool:
    return is_holiday(date_iso) or is_weekend(date_iso)

def _ok_person_on_date(rid: str, rank: str, duty: str, date_iso: str, counters: dict,
                       weekend_cap: bool = True, holiday_cap: bool = True, weekday_only: bool = False) -> bool:
    if weekday_only and not is_weekday(date_iso):
        return False

    ceiling = MAX_PER_MONTH.get(str(rank).strip(), 99) # Using 99 as a high default
    tot = counters.get(rid, {}).get("total", 0)
    if tot >= ceiling:
        return False

    is_wknd = is_weekend(date_iso)
    is_hol  = is_holiday(date_iso)
    wknd_cnt = counters.get(rid, {}).get("wknd", 0)
    hol_cnt  = counters.get(rid, {}).get("hol_real", 0)

    if weekend_cap and is_wknd and wknd_cnt >= 2:
        return False
    if holiday_cap and is_hol and hol_cnt >= 1:
        return False

    prev_dates = counters.get(rid, {}).get("dates", set())
    if not two_day_gap_ok(prev_dates, date_iso):
        return False

    return True

# ------------------------------ Core scheduler -------------------------------

def make_month_schedule_all(year: int, month: int) -> dict:
    print("\n" + "="*50)
    print(f"DEBUG: STARTING SHIFT CALCULATION for {year}-{month:02d}")
    print("="*50)

    people_map = _load_people_map()
    print(f"DEBUG: Loaded {len(people_map)} people from personnel.csv.")
    if not people_map:
        print("DEBUG: ERROR - personnel.csv is empty or could not be loaded. Aborting.")
        return {"dates": [], "by_watch": {}, "counters": {}}

    path = Path("logs") / f"ship_status_{year:04d}_{month:02d}.csv"
    print(f"DEBUG: Attempting to load ship status from: {path}")
    if not path.exists():
        print(f"DEBUG: ERROR - Ship status file not found. Aborting.")
        return {"dates": [], "by_watch": {w: {} for w in WATCH_TYPES}, "counters": {}}

    ship = pd.read_csv(path, dtype=str, encoding='utf-8-sig').fillna("")
    dates = sorted([d for d in ship["date"].astype(str).tolist() if d])
    print(f"DEBUG: Found {len(dates)} days to process in ship status file.")

    by_watch = {w: {} for w in WATCH_TYPES}
    counters = {}

    for date_iso in dates:
        print(f"\n----- PROCESSING DATE: {date_iso} -----")
        status = ship.loc[ship["date"] == date_iso, "status"].iloc[0]
        print(f"DEBUG: Ship status is '{status}'.")
        if status.lower().strip() != "in port":
            print("DEBUG: Status is not 'in port'. Marking as SEA and skipping.")
            for w in WATCH_TYPES:
                by_watch[w][date_iso] = "SEA"
            continue

        used_today = set()
        info = day_availability(date_iso)
        
        print("DEBUG: Initial pools from day_availability():")
        for w, p in info.get("pools", {}).items():
            print(f"  - Pool '{w}': {len(p)} people.")

        # AF Watch
        w_af = "ΑΦ"
        pool_af_all = _primary_pool(info, w_af, people_map, af_mode=True)
        p_weekday_only = [p for p in pool_af_all if str(p.get("duty", "")).strip() in DUTY_WEEKDAY_ONLY]
        p_regular      = [p for p in pool_af_all if str(p.get("duty", "")).strip() not in DUTY_WEEKDAY_ONLY]
        picked_af = None

        for p in _sort_af_youngest_first(p_regular):
            rid, rk = p["registry_id"], str(p["rank"]).strip()
            if rid not in used_today and _ok_person_on_date(rid, rk, "", date_iso, counters, weekday_only=False):
                picked_af = p
                break
        
        if not picked_af:
            for p in _sort_af_youngest_first(p_weekday_only):
                rid, rk = p["registry_id"], str(p["rank"]).strip()
                if rid not in used_today and _ok_person_on_date(rid, rk, "weekday_only", date_iso, counters, weekday_only=True):
                    picked_af = p
                    break

        by_watch[w_af][date_iso] = picked_af
        if picked_af:
            print(f"DEBUG: ==> Assigned {w_af}: {picked_af.get('name')}")
            rid = picked_af["registry_id"]
            c = counters.setdefault(rid, {"name": picked_af["name"], "rank": picked_af["rank"], "specialty": picked_af["specialty"], "total":0, "wknd":0, "hol":0, "hol_real":0, "dates":set(), "per_watch":{w:0 for w in WATCH_TYPES}, "hol_watch":{w:0 for w in WATCH_TYPES}})
            c["total"] += 1; c["per_watch"][w_af] += 1; c["dates"].add(date_iso)
            if is_weekend(date_iso): c["wknd"] += 1
            if _is_holiday_like(date_iso): c["hol"] += 1; c["hol_watch"][w_af] += 1
            if is_holiday(date_iso): c["hol_real"] += 1
            used_today.add(rid)
        else:
            print(f"DEBUG: ==> No one assigned to {w_af}.")

        # Other watches
        for w in ["ΥΦ", "ΥΦΜ", "ΒΥΦΜ", "ΒΥΦ"]:
            pool_w = _primary_pool(info, w, people_map)
            picked = None
            for p in _sort_fair_non_af(pool_w, counters):
                rid, rk = p["registry_id"], str(p["rank"]).strip()
                if rid not in used_today and _ok_person_on_date(rid, rk, "", date_iso, counters):
                    picked = p
                    break
            by_watch[w][date_iso] = picked
            if picked:
                print(f"DEBUG: ==> Assigned {w}: {picked.get('name')}")
                rid = picked["registry_id"]
                c = counters.setdefault(rid, {"name": picked["name"], "rank": picked["rank"], "specialty": picked["specialty"], "total":0, "wknd":0, "hol":0, "hol_real":0, "dates":set(), "per_watch":{wt:0 for wt in WATCH_TYPES}, "hol_watch":{wt:0 for wt in WATCH_TYPES}})
                c["total"] += 1; c["per_watch"][w] += 1; c["dates"].add(date_iso)
                if is_weekend(date_iso): c["wknd"] += 1
                if _is_holiday_like(date_iso): c["hol"] += 1; c["hol_watch"][w] += 1
                if is_holiday(date_iso): c["hol_real"] += 1
                used_today.add(rid)
            else:
                print(f"DEBUG: ==> No one assigned to {w}.")

    print("\n" + "="*50)
    print(f"DEBUG: CALCULATION COMPLETE. Total personnel with assignments: {len(counters)}")
    print("="*50 + "\n")
    return {"dates": dates, "by_watch": by_watch, "counters": counters}

# ------------------------------ Excel exporting ------------------------------
# (No changes to export logic)
def _apply_table_borders(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border

def export_month_schedule_all(year: int, month: int, result=None):
    if result is None:
        result = make_month_schedule_all(year, month)
    dates    = result.get("dates", [])
    by_watch = result.get("by_watch", {w: {} for w in WATCH_TYPES})
    counters = result.get("counters", {})
    outdir = Path("data")
    outdir.mkdir(parents=True, exist_ok=True)
    path_watches  = outdir / f"Shifts_{year:04d}-{month:02d}.xlsx"
    path_calendar = outdir / f"Calendar_{year:04d}-{month:02d}.xlsx"
    path_summary  = outdir / f"Monthly_Summary_{year:04d}-{month:02d}.xlsx"
    gray = PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type="solid")
    header_map = {"ΑΦ": "AF", "ΥΦ": "YF", "ΥΦΜ": "YFM", "ΒΥΦΜ": "BYFM", "ΒΥΦ": "BYF"}

    with pd.ExcelWriter(path_watches, engine="openpyxl") as xw:
        for w in WATCH_TYPES:
            rows = []
            english_watch_header = header_map.get(w, w)
            for d in dates:
                dayname = weekday_name_gr(d)
                entry = by_watch.get(w, {}).get(d)
                if entry == "SEA": name = "At Sea"
                elif entry is None: name = ""
                else: name = _display_name(entry["rank"], entry["specialty"], entry["name"])
                rows.append({"Date": d, "Day": dayname, english_watch_header: name})
            df = pd.DataFrame(rows)
            df.to_excel(xw, sheet_name=english_watch_header, index=False)
            ws = xw.book[english_watch_header]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for r in range(2, ws.max_row + 1):
                date_iso = ws[f"A{r}"].value
                if date_iso and (is_weekend(date_iso) or is_holiday(date_iso)):
                    for c in range(1, ws.max_column + 1):
                        ws[f"{get_column_letter(c)}{r}"].fill = gray
            _apply_table_borders(ws)
            recs = []
            for rid, st in counters.items():
                cnt = st.get("per_watch", {}).get(w, 0)
                hol = st.get("hol_watch", {}).get(w, 0)
                if cnt == 0 and hol == 0: continue
                recs.append({
                    "Registry No.": rid, "Full Name": st.get("name", ""),
                    "Rank": st.get("rank", ""), "Specialty": st.get("specialty", ""),
                    f"Total {english_watch_header}": cnt,
                    f"{english_watch_header} on holiday/weekend": hol,
                })
            df_sum = pd.DataFrame(recs)
            if not df_sum.empty:
                df_sum = df_sum.sort_values([f"Total {english_watch_header}", "Rank", "Full Name"], ascending=[False, True, True])
                start_col = ws.max_column + 2
                for j, h in enumerate(df_sum.columns, start=start_col): ws.cell(row=1, column=j, value=h)
                for i, row in enumerate(df_sum.itertuples(index=False), start=2):
                    for j, val in enumerate(row, start=start_col): ws.cell(row=i, column=j, value=val)
                _apply_table_borders(ws)
        recs_all = []
        for rid, st in counters.items():
            row = {"Registry No.": rid, "Full Name": st.get("name", ""), "Rank": st.get("rank", ""), "Specialty": st.get("specialty", "")}
            total = 0
            for w in WATCH_TYPES:
                c = st.get("per_watch", {}).get(w, 0)
                row[header_map.get(w, w)] = c
                total += c
            row["Total"] = total
            row["Total Holidays/Weekends"] = st.get("hol", 0)
            recs_all.append(row)
        df_all_cols = ["Registry No.", "Full Name", "Rank", "Specialty"] + [header_map.get(w, w) for w in WATCH_TYPES] + ["Total", "Total Holidays/Weekends"]
        df_all = pd.DataFrame(recs_all, columns=df_all_cols)
        if not df_all.empty:
            df_all = df_all.sort_values(["Total", "Rank", "Full Name"], ascending=[False, True, True])
        df_all.to_excel(xw, sheet_name="Statistics (Total)", index=False)
        ws = xw.book["Statistics (Total)"]
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        _apply_table_borders(ws)

    with pd.ExcelWriter(path_calendar, engine="openpyxl") as xw:
        rows = []
        for d in dates:
            rec = {"Date": d, "Day": weekday_name_gr(d)}
            for w in WATCH_TYPES:
                entry = by_watch.get(w, {}).get(d)
                if entry == "SEA": name = "At Sea"
                elif entry is None: name = ""
                else: name = _display_name(entry["rank"], entry["specialty"], entry["name"])
                rec[header_map.get(w, w)] = name
            rows.append(rec)
        df_cols = ["Date", "Day"] + [header_map.get(w, w) for w in WATCH_TYPES]
        df = pd.DataFrame(rows, columns=df_cols)
        df.to_excel(xw, sheet_name="Calendar", index=False)
        ws = xw.book["Calendar"]
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for r in range(2, ws.max_row + 1):
            date_iso = ws[f"A{r}"].value
            if date_iso and (is_weekend(date_iso) or is_holiday(date_iso)):
                for c in range(1, ws.max_column + 1):
                    ws[f"{get_column_letter(c)}{r}"].fill = PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type="solid")
        _apply_table_borders(ws)

    with pd.ExcelWriter(path_summary, engine="openpyxl") as xw:
        recs = []
        for rid, st in counters.items():
            row = {"Registry No.": rid, "Full Name": st.get("name", ""), "Rank": st.get("rank", ""), "Specialty": st.get("specialty", "")}
            total = 0
            for w in WATCH_TYPES:
                c = st.get("per_watch", {}).get(w, 0)
                row[header_map.get(w, w)] = c
                total += c
            row["Total"] = total
            row["Total Holidays/Weekends"] = st.get("hol", 0)
            recs.append(row)
        df_cols_summary = ["Registry No.", "Full Name", "Rank", "Specialty"] + [header_map.get(w, w) for w in WATCH_TYPES] + ["Total", "Total Holidays/Weekends"]
        df = pd.DataFrame(recs, columns=df_cols_summary)
        if not df.empty:
            df = df.sort_values(["Total", "Rank", "Full Name"], ascending=[False, True, True])
        df.to_excel(xw, sheet_name="Summary", index=False)
        ws = xw.book["Summary"]
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        _apply_table_borders(ws)

    return [path_watches, path_calendar, path_summary]



'''# app/scheduler_in_port.py
# -----------------------------------------------------------------------------
# In-port scheduler for all watch types (ΑΦ, ΥΦ, ΥΦΜ, ΒΥΦΜ, ΒΥΦ).
# Constraints (from your spec):
# - Captain ('Κυβερνήτης') never scheduled.
# - Max duties per month by rank (global across all watches).
# - Only PRIMARY shift considered (per watch).
# - Two-day gap between any two duties for the same person.
# - Max 2 weekend duties/month/person; max 1 holiday duty/month/person.
# - AF: assign youngest→oldest; roles Ύπαρχος/ΔΠΕ only Mon-Fri and placed after others.
# - Other watches: same constraints (no special treatment for Ύπαρχος/ΔΠΕ).
#
# Output Excel:
# - One sheet per watch type with columns: Ημερομηνία | Ημέρα | <watch>
# - Gray rows for Sat/Sun and holidays; 'Ως εν πλω' for at-sea days.
# - On the right of each sheet: per-person summary for that watch (total + holidays).
# - Final sheet "Στατιστικά (Σύνολο)" with totals across all watches.
# -----------------------------------------------------------------------------

from __future__ import annotations
from .i18n_display_mapping import I18N
from .constants import RANKS, SPECIALTIES, DUTIES
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .scheduling_prep import day_availability, WATCH_TYPES, _seniority_key
from .i18n_display_mapping import I18N
from .constants import RANKS, SPECIALTIES, DUTIES
from .scheduler_rules import (
    MAX_PER_MONTH, DUTY_NEVER, DUTY_WEEKDAY_ONLY,
    is_weekend, is_weekday, is_holiday, weekday_name_gr, two_day_gap_ok
)

# ------------------------- Name formatting rules ----------------------------

I18N_SCH = I18N(RANKS, SPECIALTIES, DUTIES, ["", "ΑΦ","ΥΦ","ΥΦΜ","ΒΥΦΜ","ΒΥΦ"]) 

def _is_officer(rank: str) -> bool:
    officers = {
        "Πλωτάρχης","Πλωτάρχης (Μ)","Υποπλοίαρχος","Υποπλοίαρχος (Μ)",
        "Ανθυποπλοίαρχος","Ανθυποπλοίαρχος (Μ)","Ανθυποπλοίαρχος (Ε)",
        "Σημαιοφόρος","Σημαιοφόρος (Μ)","Σημαιοφόρος (Ε)"
    }
    return str(rank).strip() in officers

def _display_name(rank: str, specialty: str, name: str) -> str:
    """Officers: 'Βαθμός | Όνομα Επώνυμο ΠΝ' ; Others: 'Βαθμός (Ειδ) | Όνομα Επώνυμο'."""
    rank = I18N_SCH.to_storage('rank', str(rank).strip()); specialty = I18N_SCH.to_storage('specialty', str(specialty).strip()); name = str(name).strip()
    if _is_officer(rank):
        return f"{rank} | {name} ΠΝ"
    spec = f" ({specialty})" if specialty else ""
    return f"{rank}{spec} | {name}"

# ----------------------------- Core scheduling ------------------------------

def _load_people_map() -> dict:
    """Return {registry_number: row_dict} for personnel (plain dict, not Series)."""
    people = pd.read_csv("data/personnel.csv", dtype=str).fillna("")
    out = {}
    for _, r in people.iterrows():
        key = str(r.get("registry_number", "")).strip()
        if key:
            out[key] = r.to_dict()
    return out

def _primary_pool(info: dict, watch: str, people_map: dict, af_mode=False) -> list[dict]:
    """
    From day_availability pool for a given watch, keep only those whose primary_shift == watch,
    exclude 'Κυβερνήτης'. For AF we annotate duty for weekday-only rule.
    """
    base = info["pools"].get(watch, [])
    out = []
    for p in base:
        rid = p["registry_id"]
        row = people_map.get(str(rid))
        if row is None:
            continue
        duty = str(row.get("duty","")).strip()
        if duty in DUTY_NEVER:
            continue
        if I18N_SCH.to_storage("watch", str(row.get("primary_shift","")).strip()) != watch:
            continue
        q = dict(p)
        if af_mode:
            q["duty"] = duty
        out.append(q)
    return out

def _sort_af_youngest_first(pool: list[dict]) -> list[dict]:
    """Youngest (least senior) first: reverse sort by seniority key, then by name."""
    return sorted(pool, key=lambda p: (_seniority_key(p["rank"]), str(p["name"])), reverse=True)

def _sort_fair_non_af(pool: list[dict], counters: dict) -> list[dict]:
    """
    For non-AF watches: fair ordering — least total duties, then seniority, then name.
    """
    def key(p):
        rid = p["registry_id"]
        tot = counters.get(rid, {}).get("total", 0)
        return (tot, _seniority_key(p["rank"]), str(p["name"]))
    return sorted(pool, key=key)

def _ok_person_on_date(rid: str, rank: str, duty: str, date_iso: str, counters: dict,
                       weekend_cap=True, holiday_cap=True, weekday_only=False) -> bool:
    """
    Check all generic constraints for a person on a given date.
    """
    # Weekday-only requirement (for AF with Ύπαρχος/ΔΠΕ)
    if weekday_only and not is_weekday(date_iso):
        return False

    # Monthly max per rank (global across all watches)
    ceiling = MAX_PER_MONTH.get(str(rank).strip(), 0)
    tot = counters.get(rid, {}).get("total", 0)
    if tot >= ceiling:
        return False

    # Weekend/holiday caps (global across all watches)
    is_wknd = is_weekend(date_iso)
    is_hol  = is_holiday(date_iso)
    wknd_cnt = counters.get(rid, {}).get("wknd", 0)
    hol_cnt  = counters.get(rid, {}).get("hol", 0)
    if weekend_cap and is_wknd and wknd_cnt >= 2:
        return False
    if holiday_cap and is_hol and hol_cnt >= 1:
        return False

    # Two-day gap (global)
    prev_dates = counters.get(rid, {}).get("dates", set())
    if not two_day_gap_ok(prev_dates, date_iso):
        return False

    return True

def make_month_schedule_all(year: int, month: int) -> dict:
    """
    Build assignments for all watch types for the given month.
    Returns:
      {
        "dates": [YYYY-MM-DD...],
        "by_watch": { watch: { date: person-dict|None|"SEA" } },
        "counters": { rid: {name,rank,specialty,total,wknd,hol,dates:set, per_watch:{}, hol_watch:{} } }
      }
    """
    people_map = _load_people_map()

    # Read ship status
    path = Path("logs") / f"ship_status_{year:04d}_{month:02d}.csv"
    if not path.exists():
        return {"dates": [], "by_watch": {w:{} for w in WATCH_TYPES}, "counters": {}}
    ship = pd.read_csv(path, dtype=str).fillna("")
    dates = []
    for d in ship["date"].astype(str).tolist():
        try:
            datetime.fromisoformat(d)  # validate
            dates.append(d)
        except Exception:
            continue
    dates = sorted(dates, key=lambda s: datetime.fromisoformat(s))

    # Initialize structures
    by_watch = {w: {} for w in WATCH_TYPES}
    counters = {}  # rid -> stats dict

    # Iterate days
    for date_iso in dates:
        status = ship.loc[ship["date"] == date_iso, "status"].iloc[0]
        if status != "ορμώ":
            for w in WATCH_TYPES:
                by_watch[w][date_iso] = "SEA"
            continue

        # Same-day uniqueness across all watches
        used_today = set()

        # --- 1) AF with special rules ---
        info = day_availability(date_iso)
        pool_af_all = _primary_pool(info, "ΑΦ", people_map, af_mode=True)

        # Partition AF pool into weekday-only (Ύπαρχος/ΔΠΕ) and regular
        p_weekday_only = [p for p in pool_af_all if str(p.get("duty","")).strip() in DUTY_WEEKDAY_ONLY]
        p_regular      = [p for p in pool_af_all if str(p.get("duty","")).strip() not in DUTY_WEEKDAY_ONLY]

        picked_af = None

        # First pass: regular (youngest→oldest)
        for p in _sort_af_youngest_first(p_regular):
            rid = p["registry_id"]
            rk  = str(p["rank"]).strip()
            if _ok_person_on_date(rid, rk, "", date_iso, counters,
                                  weekend_cap=True, holiday_cap=True, weekday_only=False) and rid not in used_today:
                picked_af = p
                break

        # Second pass: weekday-only roles (Mon-Fri only), youngest→oldest
        if not picked_af:
            for p in _sort_af_youngest_first(p_weekday_only):
                rid = p["registry_id"]
                rk  = str(p["rank"]).strip()
                if _ok_person_on_date(rid, rk, "weekday_only", date_iso, counters,
                                      weekend_cap=True, holiday_cap=True, weekday_only=True) and rid not in used_today:
                    picked_af = p
                    break

        by_watch["ΑΦ"][date_iso] = picked_af

        # Update counters if assigned
        if picked_af:
            rid = picked_af["registry_id"]
            c = counters.get(rid, {"name":picked_af["name"],"rank":picked_af["rank"],"specialty":picked_af["specialty"],
                                   "total":0,"wknd":0,"hol":0,"dates":set(),
                                   "per_watch":{w:0 for w in WATCH_TYPES},
                                   "hol_watch":{w:0 for w in WATCH_TYPES}})
            c["total"] += 1
            c["per_watch"]["ΑΦ"] += 1
            if is_weekend(date_iso):
                c["wknd"] += 1
            if is_holiday(date_iso):
                c["hol"] += 1
                c["hol_watch"]["ΑΦ"] += 1
            ds = set(c["dates"]); ds.add(date_iso); c["dates"] = ds
            counters[rid] = c
            used_today.add(rid)

        # --- 2) Other watches with fair ordering ---
        for w in ["ΥΦ", "ΥΦΜ", "ΒΥΦΜ", "ΒΥΦ"]:
            pool_w = _primary_pool(info, w, people_map)
            if not pool_w:
                by_watch[w][date_iso] = None
                continue

            # Fair ordering: least total → seniority → name
            picked = None
            for p in _sort_fair_non_af(pool_w, counters):
                rid = p["registry_id"]
                rk  = str(p["rank"]).strip()
                if rid in used_today:
                    continue
                if _ok_person_on_date(rid, rk, "", date_iso, counters,
                                      weekend_cap=True, holiday_cap=True, weekday_only=False):
                    picked = p
                    break

            by_watch[w][date_iso] = picked

            # Update counters
            if picked:
                rid = picked["registry_id"]
                c = counters.get(rid, {"name":picked["name"],"rank":picked["rank"],"specialty":picked["specialty"],
                                       "total":0,"wknd":0,"hol":0,"dates":set(),
                                       "per_watch":{wt:0 for wt in WATCH_TYPES},
                                       "hol_watch":{wt:0 for wt in WATCH_TYPES}})
                c["total"] += 1
                c["per_watch"][w] += 1
                if is_weekend(date_iso):
                    c["wknd"] += 1
                if is_holiday(date_iso):
                    c["hol"] += 1
                    c["hol_watch"][w] += 1
                ds = set(c["dates"]); ds.add(date_iso); c["dates"] = ds
                counters[rid] = c
                used_today.add(rid)

    return {"dates": dates, "by_watch": by_watch, "counters": counters}

# ------------------------------ Excel export --------------------------------

def export_month_schedule_all(year: int, month: int, result=None) -> Path:
    """
    Δημιουργεί ΕΝΑ Excel:
      - Ένα φύλλο ανά βάρδια (ΑΦ, ΥΦ, ΥΦΜ, ΒΥΦΜ, ΒΥΦ) με στήλες: Ημερομηνία | Ημέρα | <watch>
        * Γκρι σειρά για Σ/Κ και αργίες
        * 'Ως εν πλω' για μέρες EN_PLO
        * Μαύρα λεπτά borders σε όλο τον πίνακα
        * Δεξιά σύνοψη ανά φύλλο με σύνολα & 'σε αργία' + borders
      - Τελικό φύλλο "Στατιστικά (Σύνολο)" με αθροιστικά + borders

    Επιστρέφει το Path του αρχείου (π.χ. data/Shifts_2025-09.xlsx).
    """
    # Αν δεν περάστηκε result από έξω, τρέξε scheduler τώρα
    if result is None:
        result = make_month_schedule_all(year, month)

    dates     = result.get("dates", [])
    by_watch  = result.get("by_watch", {w: {} for w in WATCH_TYPES})
    counters  = result.get("counters", {})

    # Προορισμός
    outdir = Path("data")
    outdir.mkdir(parents=True, exist_ok=True)
    outfile = outdir / f"Shifts_{year:04d}-{month:02d}.xlsx"

    gray   = PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type="solid")
    thin   = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    with pd.ExcelWriter(outfile, engine="openpyxl") as xw:
        # ---- Φύλλα ανά βάρδια ----
        for w in WATCH_TYPES:
            rows = []
            for d in dates:
                dayname = weekday_name_gr(d)
                entry = by_watch.get(w, {}).get(d)
                if entry == "SEA":
                    name = "Ως εν πλω"
                elif entry is None:
                    name = ""
                else:
                    name = _display_name(entry["rank"], entry["specialty"], entry["name"])
                rows.append({"Ημερομηνία": d, "Ημέρα": dayname, w: name})

            df = pd.DataFrame(rows, columns=["Ημερομηνία", "Ημέρα", w])
            df.to_excel(xw, sheet_name=w, index=False)
            ws = xw.book[w]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # Γκρι για Σ/Κ & αργίες + borders σε όλα τα κελιά του κύριου πίνακα
            max_row = ws.max_row
            max_col = ws.max_column
            for r in range(2, max_row + 1):
                date_iso = ws[f"A{r}"].value
                if date_iso:
                    if is_weekend(date_iso) or is_holiday(date_iso):
                        for c in range(1, max_col + 1):
                            ws[f"{get_column_letter(c)}{r}"].fill = gray
                # borders για κάθε γραμμή
                for c in range(1, max_col + 1):
                    ws[f"{get_column_letter(c)}{r}"].border = border
            # borders και στα headers
            for c in range(1, max_col + 1):
                ws[f"{get_column_letter(c)}1"].border = border

            # ---- Δεξιά σύνοψη για τη συγκεκριμένη βάρδια ----
            recs = []
            for rid, st in counters.items():
                cnt = st.get("per_watch", {}).get(w, 0)
                hol = st.get("hol_watch", {}).get(w, 0)
                if cnt == 0 and hol == 0:
                    continue
                recs.append({
                    "Αρ. Μητρώου": rid,
                    "Ονοματεπώνυμο": st.get("name", ""),
                    "Βαθμός": st.get("rank", ""),
                    "Ειδικότητα": st.get("specialty", ""),
                    f"Σύνολο {w}": cnt,
                    f"{w} σε αργία": hol
                })
            df_sum = pd.DataFrame(
                recs,
                columns=["Αρ. Μητρώου","Ονοματεπώνυμο","Βαθμός","Ειδικότητα",f"Σύνολο {w}",f"{w} σε αργία"]
            )
            if not df_sum.empty:
                df_sum = df_sum.sort_values([f"Σύνολο {w}","Βαθμός","Ονοματεπώνυμο"], ascending=[False, True, True])

                # Τοποθέτηση δεξιά από τον κύριο πίνακα με ένα κενό
                start_col = ws.max_column + 2
                # headers
                for j, h in enumerate(df_sum.columns, start=start_col):
                    cell = ws.cell(row=1, column=j, value=h)
                    cell.border = border
                # rows
                for i, row in enumerate(df_sum.itertuples(index=False), start=2):
                    for j, val in enumerate(row, start=start_col):
                        cell = ws.cell(row=i, column=j, value=val)
                        cell.border = border

        # ---- Τελικό: "Στατιστικά (Σύνολο)" ----
        recs_all = []
        for rid, st in counters.items():
            row = {
                "Αρ. Μητρώου": rid,
                "Ονοματεπώνυμο": st.get("name",""),
                "Βαθμός": st.get("rank",""),
                "Ειδικότητα": st.get("specialty",""),
            }
            total = 0
            for w in WATCH_TYPES:
                c = st.get("per_watch", {}).get(w, 0)
                row[w] = c
                total += c
            row["Σύνολο"] = total
            row["Σύνολο Αργίες"] = st.get("hol", 0)
            recs_all.append(row)

        df_all = pd.DataFrame(
            recs_all,
            columns=["Αρ. Μητρώου","Ονοματεπώνυμο","Βαθμός","Ειδικότητα"] + WATCH_TYPES + ["Σύνολο","Σύνολο Αργίες"]
        )
        if not df_all.empty:
            df_all = df_all.sort_values(["Σύνολο","Βαθμός","Ονοματεπώνυμο"], ascending=[False, True, True])

        df_all.to_excel(xw, sheet_name="Στατιστικά (Σύνολο)", index=False)
        ws = xw.book["Στατιστικά (Σύνολο)"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        # borders σε όλο το φύλλο
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws[f"{get_column_letter(c)}{r}"].border = border

    return outfile'''